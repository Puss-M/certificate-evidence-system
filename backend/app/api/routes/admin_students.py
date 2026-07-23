"""
学生 Excel 批量导入接口（接口规范.md第4.2节"导入学生名单"）。

Excel 格式要求：第一行是表头，支持中英文两种写法（可以混用）：
    学号 / student_no      —— 必填
    姓名 / student_name    —— 必填
    学院 / college         —— 选填
    班级 / class_name      —— 选填
    专业 / major_name      —— 选填
从第二行开始每一行是一个学生。学号重复（不管是和数据库里已有的学生重复，还是
同一份表格内部自己重复）都记成失败，不会覆盖已有学生；每一行独立处理，
成功的照样落库，失败的记下具体第几行、什么原因，不会因为某一行有问题就让
整个导入失败。
"""
import io
import secrets

from fastapi import APIRouter, Depends, File, Form, HTTPException, UploadFile
from openpyxl import load_workbook
from pydantic import BaseModel, Field
from sqlalchemy.orm import Session

from app.api.routes.auth import PASSWORD_HASH, require_admin_access
from app.api.routes.certificate_batches import create_batch_record
from app.core.responses import ApiResponse
from app.db.session import get_db
from app.models.audit_log import AuditLog
from app.models.student import Student
from app.models.user import AuthSession, User
from app.schemas.student import ImportFailure, ImportResult

router = APIRouter(prefix="/admin/students", dependencies=[Depends(require_admin_access)])

_HEADER_ALIASES: dict[str, set[str]] = {
    "student_no": {"student_no", "学号"},
    "student_name": {"student_name", "姓名"},
    "college": {"college", "学院"},
    "class_name": {"class_name", "班级"},
    "major_name": {"major_name", "专业"},
}


class StudentAccountProvisionRequest(BaseModel):
    student_ids: list[int] = Field(default_factory=list, max_length=500)


def _initial_password() -> str:
    return secrets.token_urlsafe(12)


def _revoke_sessions(db: Session, user_id: int) -> None:
    from datetime import datetime

    db.query(AuthSession).filter(
        AuthSession.user_id == user_id,
        AuthSession.revoked_at.is_(None),
    ).update({AuthSession.revoked_at: datetime.utcnow()}, synchronize_session=False)


def _credential_record(student: Student, password: str) -> dict:
    return {
        "student_id": student.student_id,
        "student_no": student.student_no,
        "student_name": student.student_name,
        "initial_password": password,
    }


def _map_headers(header_row: tuple) -> dict[str, int]:
    column_index: dict[str, int] = {}
    for idx, cell in enumerate(header_row):
        if cell is None:
            continue
        cell_text = str(cell).strip()
        for field, aliases in _HEADER_ALIASES.items():
            if cell_text in aliases:
                column_index[field] = idx
    return column_index


@router.post("/accounts/provision")
def provision_student_accounts(
    payload: StudentAccountProvisionRequest,
    db: Session = Depends(get_db),
    current_user: dict = Depends(require_admin_access),
) -> ApiResponse[dict]:
    query = db.query(Student).order_by(Student.student_id.asc())
    requested_ids = set(payload.student_ids)
    if requested_ids:
        query = query.filter(Student.student_id.in_(requested_ids))
    students = query.all()
    found_ids = {student.student_id for student in students}
    missing_ids = sorted(requested_ids - found_ids)

    existing_by_student_id = {
        user.student_id: user
        for user in db.query(User).filter(User.student_id.is_not(None)).all()
        if user.student_id is not None
    }
    usernames = {student.student_no for student in students}
    occupied_usernames = {
        user.username for user in db.query(User).filter(User.username.in_(usernames)).all()
    }
    created: list[dict] = []
    skipped: list[dict] = [{"student_id": student_id, "reason": "学生不存在"} for student_id in missing_ids]

    for student in students:
        if student.student_id in existing_by_student_id:
            skipped.append({"student_id": student.student_id, "reason": "学生账号已开通"})
            continue
        if student.student_no in occupied_usernames:
            skipped.append({"student_id": student.student_id, "reason": "学号已被其他账号占用"})
            continue
        password = _initial_password()
        db.add(
            User(
                username=student.student_no,
                display_name=student.student_name,
                password_hash=PASSWORD_HASH.hash(password),
                role="STUDENT",
                student_id=student.student_id,
                must_change_password=True,
            )
        )
        created.append(_credential_record(student, password))

    if created:
        db.add(
            AuditLog(
                action="批量开通学生账号",
                target_type="学生账号",
                target_id=None,
                operator=current_user["username"],
                detail=f"created={len(created)}, skipped={len(skipped)}",
            )
        )
        db.commit()
    return ApiResponse.success({"created": created, "skipped": skipped})


@router.post("/{student_id}/account/reset-password")
def reset_student_password(
    student_id: int,
    db: Session = Depends(get_db),
    current_user: dict = Depends(require_admin_access),
) -> ApiResponse[dict]:
    student = db.get(Student, student_id)
    if student is None:
        raise HTTPException(status_code=404, detail="学生不存在")
    user = db.query(User).filter(User.student_id == student_id, User.role == "STUDENT").one_or_none()
    if user is None:
        raise HTTPException(status_code=404, detail="该学生尚未开通账号")
    password = _initial_password()
    user.password_hash = PASSWORD_HASH.hash(password)
    user.must_change_password = True
    _revoke_sessions(db, user.user_id)
    db.add(
        AuditLog(
            action="重置学生密码",
            target_type="学生账号",
            target_id=str(student_id),
            operator=current_user["username"],
            detail="initial_password_regenerated",
        )
    )
    db.commit()
    return ApiResponse.success(_credential_record(student, password))


@router.post("/import", response_model=ApiResponse[ImportResult])
async def import_students(
    file: UploadFile = File(...),
    batch_name: str = Form(...),
    template_id: int = Form(...),
    db: Session = Depends(get_db),
) -> ApiResponse[ImportResult]:
    content = await file.read()
    try:
        workbook = load_workbook(io.BytesIO(content), data_only=True)
    except Exception as exc:  # openpyxl对损坏/非Excel文件会抛各种异常，统一按400处理
        raise HTTPException(status_code=400, detail=f"Excel文件解析失败：{exc}") from exc

    sheet = workbook.active
    rows = list(sheet.iter_rows(values_only=True))
    if not rows:
        raise HTTPException(status_code=400, detail="Excel文件是空的")

    column_index = _map_headers(rows[0])
    if "student_no" not in column_index or "student_name" not in column_index:
        raise HTTPException(
            status_code=400,
            detail="表头必须包含「学号/student_no」和「姓名/student_name」两列",
        )

    existing_student_no = {row[0] for row in db.query(Student.student_no).all()}
    failures: list[ImportFailure] = []
    imported_students: list[Student] = []

    for row_number, row in enumerate(rows[1:], start=2):  # 第1行是表头，数据从第2行开始
        raw_student_no = row[column_index["student_no"]]
        raw_student_name = row[column_index["student_name"]]

        if raw_student_no is None or str(raw_student_no).strip() == "":
            failures.append(ImportFailure(row=row_number, reason="学号为空"))
            continue
        student_no = str(raw_student_no).strip()

        if raw_student_name is None or str(raw_student_name).strip() == "":
            failures.append(ImportFailure(row=row_number, reason="姓名为空"))
            continue
        student_name = str(raw_student_name).strip()

        if student_no in existing_student_no:
            failures.append(ImportFailure(row=row_number, reason=f"学号重复：{student_no}"))
            continue

        class_name = row[column_index["class_name"]] if "class_name" in column_index else None
        major_name = row[column_index["major_name"]] if "major_name" in column_index else None
        college = row[column_index["college"]] if "college" in column_index else None
        college_text = str(college).strip() if college is not None else None
        if college_text and len(college_text) > 100:
            failures.append(ImportFailure(row=row_number, reason="学院名称不能超过100个字符"))
            continue

        student = Student(
            student_no=student_no,
            student_name=student_name,
            college=college_text or None,
            class_name=str(class_name).strip() if class_name else None,
            major_name=str(major_name).strip() if major_name else None,
        )
        db.add(student)
        imported_students.append(student)
        existing_student_no.add(student_no)  # 防止同一份表格内部学号自己重复

    db.commit()

    # 导入成功后顺带建一个批次，把这次成功导入的学生直接作为这个批次的student_ids
    # ——对应接口规范.md第4.3节要求创建批次时就要有名单。这不是唯一建批次的方式，
    # 管理员之后也可以在"批次管理"页面用POST /admin/batches手动挑学生单独建批次。
    create_batch_record(
        db,
        batch_name=batch_name,
        template_id=template_id,
        student_ids=[student.student_id for student in imported_students],
    )
    success_count = len(imported_students)

    return ApiResponse.success(
        ImportResult(
            success_count=success_count,
            failed_count=len(failures),
            failures=failures,
        )
    )
